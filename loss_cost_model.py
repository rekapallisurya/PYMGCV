# ==============================================================================
# Insurance Loss Cost Model — pymgcv GAM (Tweedie, REML)
# Mirrors the R mgcv specification exactly
# ==============================================================================
#
# R equivalent:
#   library(mgcv)
#   model <- gam(
#     capped_claims_kusd ~ offset(log_duration) +
#       s(vehicle_age, k=8) + s(driver_age, k=10) + s(bonus_malus, k=8) +
#       s(log10_exposure_usd, k=6) + s(annual_mileage_km, k=6) +
#       class_B + class_C + class_D + is_urban + is_suburban,
#     data=df, family=Tweedie(p=1.5, link="log"), method="REML",
#     gamma=1.2, control=gam.control(maxit=150, epsilon=1e-5)
#   )
# ==============================================================================

from __future__ import annotations

import numpy as np
import pandas as pd
import openpyxl
from pymgcv import GAM, aic, plot_smooth, plot_residuals

try:
    import matplotlib.pyplot as plt
    _PLOT = True
except ImportError:
    _PLOT = False
    print("matplotlib not found — plots will be skipped. "
          "Install with: pip install matplotlib")

# ------------------------------------------------------------------------------
# 1. Load data
# ------------------------------------------------------------------------------
CSV_FILE = "insurance_loss_cost_data.csv"

df = pd.read_csv(CSV_FILE)
print(f"Rows: {len(df)} | Cols: {df.shape[1]}")
print(df.dtypes.to_string())

# ------------------------------------------------------------------------------
# 2. Fit GAM
#    Tweedie family: p = 1.5 (compound Poisson-Gamma), log link
#    gamma = 1.2  → mild over-smoothing (matches R gamma=1.2)
#    method = "REML"
#    k values match the R s() calls
# ------------------------------------------------------------------------------
formula = (
    "capped_claims_kusd ~ "
    "offset(log_duration) + "
    "s(vehicle_age, k=8) + "
    "s(driver_age, k=10) + "
    "s(bonus_malus, k=8) + "
    "s(log10_exposure_usd, k=6) + "
    "s(annual_mileage_km, k=6) + "
    "class_B + class_C + class_D + "
    "is_urban + is_suburban"
)

model = GAM(
    formula = formula,
    data    = df,
    family  = 'tweedie',          # TweedieFamily(power=1.5), log link
    method  = 'REML',
    gamma   = 1.2,                # mild over-smoothing — same as R gamma=1.2
    control = {'maxit': 150, 'epsilon': 1e-5},
)
model.fit(verbose=False)

# ------------------------------------------------------------------------------
# 3. Model summary
# ------------------------------------------------------------------------------
print("\n========== Model Summary ==========")
print(model.summary())

# ------------------------------------------------------------------------------
# 4. Smooth term plots  (mirrors R: par(mfrow=c(2,3)); plot(model, ...))
# ------------------------------------------------------------------------------
SMOOTH_VARS = [
    'vehicle_age',
    'driver_age',
    'bonus_malus',
    'log10_exposure_usd',
    'annual_mileage_km',
]

if _PLOT:
    fig, axes = plt.subplots(2, 3, figsize=(15, 8))
    axes = axes.ravel()

    for i, var in enumerate(SMOOTH_VARS):
        plot_smooth(
            model,
            var_name         = var,
            ax               = axes[i],
            n_grid           = 100,
            confidence_band  = True,   # shade=TRUE in R
            ci               = 0.95,
        )
        axes[i].set_title(f's({var})', fontsize=10)

    # Hide unused 6th panel
    axes[5].set_visible(False)

    fig.suptitle("Smooth terms — Loss Cost GAM (Tweedie, REML)", fontsize=12)
    plt.tight_layout()
    plt.savefig("smooth_terms.png", dpi=120, bbox_inches='tight')
    plt.show()
    print("Smooth term plot saved -> smooth_terms.png")

# ------------------------------------------------------------------------------
# 5. Diagnostics  (mirrors R: par(mfrow=c(2,2)); gam.check(model, k.rep=20))
# ------------------------------------------------------------------------------
check = model.gam_check(type='deviance', print_summary=True, plot=False)

if _PLOT:
    fig2, axes2 = plt.subplots(2, 2, figsize=(10, 8))
    plot_residuals(model, ax_array=axes2.ravel())
    fig2.suptitle("GAM Diagnostics — Loss Cost Model", fontsize=12)
    plt.tight_layout()
    plt.savefig("gam_diagnostics.png", dpi=120, bbox_inches='tight')
    plt.show()
    print("Diagnostic plot saved -> gam_diagnostics.png")

# ------------------------------------------------------------------------------
# 6. Key scalars  (mirrors R output after gam.check)
# ------------------------------------------------------------------------------
# Deviance and null deviance — computed on linear predictor
X      = model._X_fit
y      = model._y_fit
family = model.family

from pymgcv.utils.model_matrix import ModelMatrix as _MM
_mm     = model.model_matrix
_offset = _mm.offset_vector() if (hasattr(_mm, 'offset_vector') and _mm.offset_vector() is not None) else np.zeros(len(y))

eta_fit     = X @ model.beta + _offset
mu_fit      = family.linkinv(eta_fit)
deviance_   = float(-2 * family.loglik(y, mu_fit, dispersion=1.0))

mu_null     = np.full(len(y), y.mean())
null_dev_   = float(-2 * family.loglik(y, mu_null, dispersion=1.0))
dev_expl_   = (null_dev_ - deviance_) / null_dev_ * 100 if null_dev_ != 0 else 0.0
aic_val     = aic(model)

print(f"\n--- Tweedie power (p)  : {model.family.power}")
print(f"--- Dispersion         : {model.dispersion_:.6f}")
print(f"--- Deviance explained : {dev_expl_:.2f} %")
print(f"--- AIC                : {aic_val:.4f}")

# ------------------------------------------------------------------------------
# 7. Predictions (fitted values on response scale — kUSD)
#    Mirrors R: df$predicted_loss_cost <- predict(model, type="response")
# ------------------------------------------------------------------------------
df['predicted_loss_cost'] = model.predict(df, scale='response')

print("\nTop 6 rows — actual vs predicted (kUSD):")
print(
    df[['policy_id', 'capped_claims_kusd', 'predicted_loss_cost']]
    .head(6)
    .to_string(index=False)
)

# Balance ratio (total predicted / total actual) — quick sanity check
total_actual = df['capped_claims_kusd'].sum()
total_pred   = df['predicted_loss_cost'].sum()
balance_ratio = total_pred / total_actual
print(f"\nBalance ratio (pred / actual): {balance_ratio:.4f}  (1.00 = perfect)")

# Pearson residuals
p_est   = model.family.power
phi_est = model.dispersion_
actual_k = df['capped_claims_kusd'].values
fitted_k = df['predicted_loss_cost'].values
pearson  = (actual_k - fitted_k) / np.sqrt(
    np.maximum(phi_est * np.maximum(fitted_k, 1e-9) ** p_est, 1e-9)
)
df['pearson_residual'] = pearson

# ------------------------------------------------------------------------------
# 8. Export to Excel — modeloutput.xlsx
# ------------------------------------------------------------------------------
EXCEL_FILE = "modeloutput.xlsx"

# ── Sheet 1: Predictions ─────────────────────────────────────────────────────
pred_cols = [
    'policy_id', 'vehicle_class', 'region',
    'vehicle_age', 'driver_age', 'bonus_malus',
    'log10_exposure_usd', 'annual_mileage_km',
    'duration', 'capped_claims_kusd',
    'predicted_loss_cost', 'pearson_residual',
]
df_pred = df[pred_cols].copy()
df_pred['predicted_loss_usd'] = df_pred['predicted_loss_cost'] * 1_000.0

# ── Sheet 2: Model scalars ────────────────────────────────────────────────────
df_scalars = pd.DataFrame({
    'Metric': [
        'Family',
        'Tweedie power (p)',
        'Link function',
        'Method',
        'Gamma (over-smoothing)',
        'Observations (n)',
        'Total EDF',
        'Parametric df',
        'Null deviance',
        'Residual deviance',
        'Deviance explained (%)',
        'AIC',
        'Dispersion (φ)',
        'Balance ratio (pred/actual)',
    ],
    'Value': [
        'Tweedie',
        model.family.power,
        'log',
        'REML',
        1.2,
        len(df),
        round(model.edf, 4),
        len(model.beta),
        round(null_dev_, 4),
        round(deviance_, 4),
        round(dev_expl_, 4),
        round(aic_val, 4),
        round(phi_est, 6),
        round(balance_ratio, 6),
    ],
})

# ── Sheet 3: Smooth term EDF ─────────────────────────────────────────────────
smooth_rows = []
for term, info in model.edf_per_smooth.items():
    smooth_rows.append({'Smooth term': term, 'EDF': round(info['edf'], 4)})
smooth_rows.append({'Smooth term': 'TOTAL', 'EDF': round(model.edf, 4)})
df_smooth = pd.DataFrame(smooth_rows)

# ── Sheet 4: Smoothing parameters ────────────────────────────────────────────
smooth_labels = list(model.edf_per_smooth.keys())
df_lambda = pd.DataFrame({
    'Smooth term': [
        smooth_labels[j] if j < len(smooth_labels) else f'penalty_{j+1}'
        for j in range(len(model.smoothing_parameters))
    ],
    'Lambda (λ)': [round(float(l), 6) for l in model.smoothing_parameters],
})

# ── Sheet 5: Parametric coefficients ─────────────────────────────────────────
from pymgcv.utils.model_matrix import ModelMatrix as _ColNames
_tmp = _ColNames(df, formula)
col_names = _tmp.column_names

PARAM_META = {
    'class_B':    'Vehicle class B vs A (medium saloon vs small)',
    'class_C':    'Vehicle class C vs A (large/performance vs small)',
    'class_D':    'Vehicle class D vs A (SUV vs small)',
    'is_urban':   'Urban vs Rural region',
    'is_suburban':'Suburban vs Rural region',
}
param_rows = []
for col, desc in PARAM_META.items():
    if col in col_names:
        idx = col_names.index(col)
        b   = float(model.beta[idx])
        eb  = float(np.exp(b))
        param_rows.append({
            'Parameter':    col,
            'Description':  desc,
            'Beta (β)':     round(b, 6),
            'exp(β)':       round(eb, 6),
            'Change %':     round((eb - 1) * 100, 2),
        })
df_params = pd.DataFrame(param_rows)

# ── Sheet 6: Segment analysis ────────────────────────────────────────────────
seg_rows = []
base_pred = df['predicted_loss_cost'].mean()
for seg_col, seg_vals in [('vehicle_class', ['A','B','C','D']),
                           ('region', ['Rural','Suburban','Urban'])]:
    for sv in seg_vals:
        mask  = df[seg_col] == sv
        n_seg = int(mask.sum())
        act   = df.loc[mask, 'capped_claims_kusd'].sum() / df.loc[mask, 'duration'].sum()
        pred  = df.loc[mask, 'predicted_loss_cost'].mean()
        seg_rows.append({
            'Segment column': seg_col,
            'Segment value':  sv,
            'N':              n_seg,
            'Actual annual loss cost (kUSD)':    round(act, 4),
            'Predicted annual loss cost (kUSD)': round(pred, 4),
            'Lift vs portfolio mean':            round(pred / base_pred, 4),
        })
df_segments = pd.DataFrame(seg_rows)

# ── Sheet 7: Residual diagnostics ────────────────────────────────────────────
df_resid_stats = pd.DataFrame({
    'Metric': [
        'Pearson residual mean',
        'Pearson residual std',
        'Pearson residual skew',
        'Pearson residual 5th pct',
        'Pearson residual 95th pct',
        'n / EDF ratio',
    ],
    'Value': [
        round(float(pearson.mean()), 6),
        round(float(pearson.std()),  6),
        round(float(pd.Series(pearson).skew()), 6),
        round(float(np.percentile(pearson, 5)),  6),
        round(float(np.percentile(pearson, 95)), 6),
        round(len(df) / model.edf, 2),
    ],
})

# ── Write workbook ────────────────────────────────────────────────────────────
HEADER_FMT = {'bold': True, 'bg_color': '#1F4E79', 'font_color': 'white',
              'border': 1, 'align': 'center', 'valign': 'vcenter'}

with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
    sheets = [
        (df_pred,        'Predictions',          'Policy-level actuals vs fitted values'),
        (df_scalars,     'Model Scalars',         'High-level model statistics'),
        (df_smooth,      'Smooth EDF',            'Effective degrees of freedom per smooth term'),
        (df_lambda,      'Smoothing Parameters',  'Fitted penalty strengths (λ)'),
        (df_params,      'Parametric Coefs',      'Parametric coefficient table with exp(β) multipliers'),
        (df_segments,    'Segment Analysis',      'Actual vs predicted by vehicle class and region'),
        (df_resid_stats, 'Residual Diagnostics',  'Pearson residual summary statistics'),
    ]

    for frame, sheet_name, description in sheets:
        frame.to_excel(writer, sheet_name=sheet_name, index=False, startrow=2)
        ws = writer.sheets[sheet_name]

        # Description row
        ws.cell(row=1, column=1, value=description)
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        ws.cell(row=1, column=1).font  = Font(italic=True, color='595959')

        # Header formatting
        thin = Side(border_style='thin', color='CCCCCC')
        for cell in ws[3]:
            cell.font      = Font(bold=True, color='FFFFFF')
            cell.fill      = PatternFill('solid', fgColor='1F4E79')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = Border(bottom=thin)

        # Auto-width columns
        for col_cells in ws.iter_cols(min_row=3, max_row=ws.max_row):
            col_letter = col_cells[0].column_letter
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0) for c in col_cells
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

        # Freeze header row
        ws.freeze_panes = 'A4'

print(f"\nExported -> {EXCEL_FILE}")
print(f"  Sheets: Predictions | Model Scalars | Smooth EDF | Smoothing Parameters")
print(f"          Parametric Coefs | Segment Analysis | Residual Diagnostics")
